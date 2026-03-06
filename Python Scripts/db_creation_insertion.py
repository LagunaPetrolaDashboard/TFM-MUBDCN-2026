
# TODO -------------------------------------SCRIPT INFORMATION-----------------------------------------------------
'''
@author: David López Lorenzo

@description: This script:
    - Creates a SQLite database file in the given directory
    - Add 3 tables to it [samples, compounds, components]
    - Insert the corresponding values of an Excel File into them.

#inputs: It takes 2 positional inputs:
    - Directory to where the Database folder that contains the SQLite.db file will be created
    - The path/directory of an Excel (xlsx) file to extract the data from

#example:
    python db_creation_insertion.py "C:\Users\example\Desktop" "C:\Users\example\Desktop\excelFile.xlsx"
'''

# TODO ----------------------------------CHECK INSTALLED MODULES---------------------------------------------------
import argparse
import os
try:
    import openpyxl
    from datetime import datetime, date
    from sqlalchemy import create_engine, ForeignKey, String, Date, Numeric, Integer, CheckConstraint, inspect
    from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker
    from sqlalchemy_utils import database_exists, create_database

except ImportError as e:
    print("\n\n" + "-" * 60)
    print("Please install the following modules to run the script:")
    print("pip install openpyxl")
    print("pip install sqlalchemy")
    print("pip install sqlalchemy-utils")
    print("\n\n" + "-" * 60)
    exit(1) # Paramos la ejecución del script



# TODO ----------------------------------DEFINED FUNCTIONS---------------------------------------------------
def normalize_dates(date_list):
    output_list = []
    dt = None
    year = None
    month = None
    for date in date_list:
        # Algunas de las fechas tienen espacios, los eliminamos
        date = date.replace(" ", "")

        # El año siempre viene defnido por los 4 ultimos caracteres y el mes por los 2 anteriores al año
        # Algunos de los dias vienen no definidos excatemente (ej: 17_19022020), nos quedamos con el último dia del rango por comodidad
        year = date[-4:]
        month = date[-6:-4]
        day = date[-8:-6]

        # Creamos un datetime y lo formateamos a YYYY-MM-DD (nos vendrá bien este formato para insertar en la base de datos)
        dt = datetime(int(year), int(month), int(day))
        output_list.append(dt.strftime("%Y-%m-%d"))

    return output_list

def main():
    # TODO ----------------------------------------INPUTS PREPARATION------------------------------------------
    # Procesamos el directorio que se ha pasado como parametro
    parser = argparse.ArgumentParser(description="Specify the directory where the SQLite database will be created and the Excel file to extract data from.")
    parser.add_argument(
        'db_directory',  # Argumento posicional (sin guion, solo el nombre)
        help="Specify the directory where the SQLite database will be created. A folder named 'Database/' will be created within this directory."
    )
    parser.add_argument(
        'xlsx_directory',  # Ruta para el archivo XLSX
        help="Provide the full path to the Excel file (.xlsx) from which the data will be extracted and inserted into the database."
    )

    args = parser.parse_args()
    ruta = args.db_directory
    xlsx = args.xlsx_directory

    # TODO --------------------------------DATABASE CREATION----------------------------------------------------
    # Si el directorio de creacion de la base de datos existe -> creamos en él la base de datos
    if os.path.exists(ruta):
        # Creamos la carpeta donde vamos a guardar el archivo (Database)
        database_dir = os.path.join(ruta, 'Database')
        if not os.path.exists(database_dir):
            os.makedirs(database_dir)

        # Creamos la ruta compuesta
        database_path = os.path.join(ruta,'Database', 'SQLite.db')
        database_url = f"sqlite:///{os.path.join(ruta, 'Database', 'SQLite.db')}"

        # Si existe ya un fichero database en esa ruta, lo eliminamos
        if os.path.exists(database_path):
            os.remove(database_path)
            print("\n-> Database eliminated")

        # Crear la base de datos de nuevo
        engine = create_engine(database_url)
        if not database_exists(engine.url):
            create_database(engine.url)
        print(f"-> Database successfully created at {database_path}")
    else:
        print(f"Error: The given directory is not valid: {ruta}")


    #TODO ----------------------------------TABLE CREATION----------------------------------------------------

    # Si el directorio de creacion de la base de datos existe -> creamos en él la base de datos
    if os.path.exists(xlsx):
        Session = sessionmaker(bind=engine)
        session = Session()

        class Base(DeclarativeBase):
            pass

        class Components(Base):
            __tablename__ = "components"

            sample_name: Mapped[str] = mapped_column(String(20), primary_key=True)

        class Compounds(Base):
            __tablename__ = "compounds"

            cas: Mapped[str] = mapped_column(String(20), primary_key=True)
            name: Mapped[str] = mapped_column(String(80))
            formula: Mapped[str] = mapped_column(String(20))
            group: Mapped[str] = mapped_column(String(30), nullable=True)

        class Samples(Base):
            __tablename__ = "samples"

            id: Mapped[int] = mapped_column(Integer, autoincrement=True, primary_key=True)
            component_id: Mapped[str] = mapped_column(ForeignKey("components.sample_name"))
            compound_id: Mapped[str] = mapped_column(ForeignKey("compounds.cas"))
            component_rt: Mapped[float] = mapped_column(
                Numeric(17, 13))  # 13 digitos en la parte decimal -> (0000.0000000000000 - 9999.9999999999999)
            library_rt: Mapped[float] = mapped_column(Numeric(17, 13), nullable=True)
            match_factor: Mapped[float] = mapped_column(
                Numeric(16, 13))  # como es un percentaje la parte entera admite solo 3 digitos y la decimal 13 digitos
            sample_date: Mapped[date] = mapped_column(Date)

            __table_args__ = (
                # No podemos insertar un porcentaje mayor a 100%
                CheckConstraint("match_factor <= 100", name="check_match_factor_max_100"),
            )

        Base.metadata.create_all(engine)

        inspector = inspect(engine)
        tables = inspector.get_table_names()

        print(f"-> Tables created: {tables}")


    # TOD0 ----------------------------------TUPLES INSERTION----------------------------------------------------
        print("-> The tuple insertion process is starting. This may take a few minutes.")

        # Cargamos el primer excel
        workbook = openpyxl.load_workbook(xlsx)
        component_tuples = 0
        compound_tuples = 0
        sample_tuples = 0
        i = 0
        total_rows = sum(sheet.max_row - 4 for sheet in workbook.worksheets)

        # Creamos las columnas Date y Group
        for sheet in workbook.worksheets:
            formated_date = datetime.strptime(normalize_dates([sheet.title])[0], "%Y-%m-%d").date()
            for row in sheet.iter_rows(min_row=5):
                i += 1
                # print(f"Row {i}/{total_rows}")
                print(f"\r\tRow {i}/{total_rows} - {((i / total_rows) * 100):.2f}%",end="")  # Sobrescribe la línea con el nuevo progreso
                # row[0] = Component RT | row[1] = Library RT 	| row[2] = Compound Name	| row[3] = Match Factor
                # row[4] = Formula	    | row[5] = CAS	        | row[6] = Sample Name
                component = Components(
                    sample_name=row[6].value)

                compound = Compounds(
                    cas=row[5].value,
                    name=row[2].value,
                    formula=row[4].value,
                    group=None)  # por el momento lo dejamos en None / NULL

                library_rt_value = row[1].value if row[1].value != '' else None
                sample = Samples(
                    component_id=component.sample_name,
                    compound_id=compound.cas,
                    component_rt=row[0].value,
                    library_rt=library_rt_value,  # row[1].value
                    match_factor=row[3].value,
                    sample_date=formated_date)

                # Verificar si el componente ya existe
                existing_component = session.query(Components).filter_by(sample_name=row[6].value).first()
                if not existing_component:
                    session.add(component)
                    component_tuples += 1

                # Verificar si el compuesto ya existe
                existing_compound = session.query(Compounds).filter_by(cas=row[5].value).first()
                if not existing_compound:
                    session.add(compound)
                    compound_tuples += 1

                # Verificar si la muestra ya existe
                existing_sample = session.query(Samples).filter_by(
                    component_id=component.sample_name,
                    compound_id=compound.cas,
                    component_rt=row[0].value,
                    library_rt=library_rt_value,  # row[1].value
                    match_factor=row[3].value,
                    sample_date=formated_date
                ).first()

                if not existing_sample:
                    session.add(sample)
                    sample_tuples += 1

                # Guardar los cambios en la base de datos
                session.commit()
        print("\n\n" + "-"*60)
        print("-> All tuples were successfully inserted into the database.")
        print(f"-> {component_tuples} tuples inserted into Components")
        print(f"-> {compound_tuples} tuples inserted into Compounds")
        print(f"-> {sample_tuples} tuples inserted into Samples")
        print("-" * 60)
    else:
        print(f"Error: The given xlsx file doesn't exist: {xlsx}")

if __name__ == "__main__":
    main()
