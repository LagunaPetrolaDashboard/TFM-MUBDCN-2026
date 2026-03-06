from sqlalchemy import create_engine
from datetime import date
from sqlalchemy import ForeignKey, String, Date, Numeric, Integer, CheckConstraint
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker
import os
from datetime import datetime
import openpyxl

# ESTE ARCHIVO ES UNA PRUEBA PARA SABER COMO SE COMPORTA LA BASE DE DATOS UNA VEZ RELLENADA CUANDO LE LLEGA UN ARCHIVO EXCEL NUEVO Y DEBE INGRESAR LOS DATOS, PARA SIMULAR QUE ESTAMOS HACIENDOLO DESDE STREAMLIT


'''
--------------------------------------------------------------------------------------
------------------------CONEXION A LA BASE DE DATOS----------------------------------
--------------------------------------------------------------------------------------
'''

# Ruta al archivo de base de datos existente
database_path = "../Database/Petrola.db"
database_url = f"sqlite:///{database_path}"

# Crear motor de conexión
engine = create_engine(database_url)
print(f"Conectado a base de datos existente en: {database_path}")



'''
--------------------------------------------------------------------------------------
----------------------------CREACIÓN DE CLASES----------------------------------------
--------------------------------------------------------------------------------------
'''
Session = sessionmaker(bind=engine)
session = Session()


class Base(DeclarativeBase):
    pass


class Stations(Base):
    __tablename__ = "stations"

    station_id: Mapped[str] = mapped_column(String, primary_key=True)
    st_type: Mapped[str] = mapped_column(String(20), nullable=True)
    geology: Mapped[str] = mapped_column(String(20), nullable=True)
    # well_depth: Mapped[float] = mapped_column(Numeric(8,3), nullable=True)
    # elevation: Mapped[float] = mapped_column(Numeric(8,3), nullable=True)
    x: Mapped[float] = mapped_column(Numeric(10, 3), nullable=False)
    y: Mapped[float] = mapped_column(Numeric(10, 3), nullable=False)


class Compounds(Base):
    __tablename__ = "compounds"

    cas: Mapped[str] = mapped_column(String(20), primary_key=True)
    name: Mapped[str] = mapped_column(String(80), nullable=False)
    formula: Mapped[str] = mapped_column(String(20), nullable=False)
    group: Mapped[str] = mapped_column(String(30), nullable=True)


class Samples(Base):
    __tablename__ = "samples"

    id: Mapped[int] = mapped_column(Integer, autoincrement=True, primary_key=True)
    station_id: Mapped[str] = mapped_column(ForeignKey("stations.station_id"), nullable=False)
    compound_cas: Mapped[str] = mapped_column(ForeignKey("compounds.cas"), nullable=False)
    component_rt: Mapped[float] = mapped_column(Numeric(17, 13),
                                                nullable=False)  # 13 digitos en la parte decimal -> (0000.0000000000000 - 9999.9999999999999)
    library_rt: Mapped[float] = mapped_column(Numeric(17, 13), nullable=True)
    match_factor: Mapped[float] = mapped_column(Numeric(16, 13),
                                                nullable=False)  # como es un percentaje la parte entera admite solo 3 digitos y la decimal 13 digitos
    sample_date: Mapped[date] = mapped_column(Date, nullable=False)

    __table_args__ = (
        # No podemos insertar un porcentaje negativo o mayor a 100%
        CheckConstraint("match_factor >= 0 AND match_factor <= 100", name="check_match_factor_range"),
    )


Base.metadata.create_all(engine)
print("Database Tables created")



'''
--------------------------------------------------------------------------------------
----------------------------INSERCCION DE NUEVAS MUESTRAS-----------------------------
--------------------------------------------------------------------------------------
'''


# Funciones auxiliares para la insercción
def sheet_is_valid(sheet):
    return not len(sheet.title) < 8


def row_is_valid(row):
    for i in range(7):
        if row[i].value not in (None, "", ''):
            return True
    return False


def excel_is_valid(excel):
    return not excel[0:2] == "GW"


# Función para poner todas las fechas en el mismo formato: 'YYYY-MM-DD'
def normalize_dates(date_list):
    output_list = []
    dt = None
    year = None
    month = None
    for date in date_list:
        try:
            # Algunas de las fechas tienen espacios, los eliminamos
            date = date.replace(" ", "")

            # El año siempre viene defnido por los 4 ultimos caracteres y el mes por los 2 anteriores al año
            # Algunos de los dias vienen no definidos excatemente (ej: 17_19022020), nos quedamos con el último dia del rango por comodidad
            year = date[-4:]
            month = date[-6:-4]
            day = date[-8:-6]

            # Creamos un datetime y lo formateamos a YYYY-MM-DD (nos vendrá bien este formato para insertar en la base de datos)
            dt = datetime(int(year), int(month), int(day))
            output_list.append(dt.strftime("%Y-%m-%d"))
        except Exception as e:
            print(f"\nError: {e}")
            print(
                f"Nombre de página incorrecto: {date_list[0]}, por favor corrija el nombre al formato DDMMYYYY en el archivo y pruebe de nuevo")

    return output_list


def get_station_date(Sample_Name):
    StationID = Sample_Name.split('_')[0]

    # HAY DOS TIPOS DE SAMPLE NAME, LOS QUE ACABAN EN SCAN Y LOS QUE NO
    # SCAN: TIENE EL FORMATO DE IDSTATION_DIA_MES_AÑO_SCAN
    # SV: TIENE EL FORMATO IDSTATION_DIAMESAÑO_FS_SV

    if len(Sample_Name.split('_')[1]) <= 2:
        day = Sample_Name.split('_')[1]
        month = Sample_Name.split('_')[2]
        year = Sample_Name.split('_')[3]
    else:
        day = Sample_Name.split('_')[1][0:2]
        month = Sample_Name.split('_')[1][2:4]
        year = '20' + Sample_Name.split('_')[1][4:6]
    dt = datetime(int(year), int(month), int(day))

    return StationID, dt.date()


# Ruta del Excel
excels_path_input = "../Datos Excel/Lecturas Pruebas Inserccion"
excel_number = 0
compound_tuples = 0
sample_tuples = 0

for excel_name in os.listdir(excels_path_input):

    # Si el nombre del excel no es valido, saltamos al siguiente
    '''if not excel_is_valid(excel_name):
        excel_number += 1
        continue'''

    # Obtenemos la ubicación del excel
    excel_path_input = os.path.join(excels_path_input, excel_name)
    excel_number += 1
    print(f"\nLoading Excel {excel_number} of {len(os.listdir(excels_path_input))}...")

    # Cargamos el primer excel
    workbook = openpyxl.load_workbook(excel_path_input)

    compound_tuples_aux = 0
    sample_tuples_aux = 0
    i = 0
    total_rows = sum(sheet.max_row - 4 for sheet in workbook.worksheets)

    # Creamos las columnas Date y Group
    for sheet in workbook.worksheets:
        # Si la hoja no es valida, pasamos a la siguiente
        if not sheet_is_valid(sheet):
            continue

        # formated_date = datetime.strptime(normalize_dates([sheet.title])[0], "%Y-%m-%d").date()

        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
            i += 1
            print(f"\r\tRow {i}/{total_rows} - {((i / total_rows) * 100):.2f}%",
                  end="")  # Sobrescribe la línea con el nuevo progreso
            # row[0] = Component RT | row[1] = Library RT 	| row[2] = Compound Name	| row[3] = Match Factor
            # row[4] = Formula	    | row[5] = CAS	        | row[6] = Sample Name      / row[7] = Compound Group

            # Antes de realizar la carga y la insercción comprobamos que la fila tiene datos
            if not row_is_valid(row):
                continue

            compound = Compounds(
                cas=row[5].value,
                name=row[2].value,
                formula=row[4].value,
                group = row[7].value if row[7].value not in (None, '') else 'Otros')

            library_rt_value = row[1].value if row[1].value != '' else None
            StationID, dt = get_station_date(row[6].value)

            if StationID == '2571b':
                StationID = '2571'

            existing_station = session.query(Stations).filter(Stations.station_id == StationID).first()
            if not existing_station:
                continue

            sample = Samples(
                station_id=StationID,
                compound_cas=row[5].value,
                component_rt=row[0].value,
                library_rt=library_rt_value,
                match_factor=row[3].value,
                sample_date=dt)
            '''print(f"{sample.station_id} - {sample.compound_cas} - {sample.component_rt} - {sample.library_rt}- {sample.match_factor} - {sample.sample_date}")'''

            # Verificar si el compuesto ya existe
            existing_compound = session.query(Compounds).filter_by(cas=row[5].value).first()
            if not existing_compound:
                session.add(compound)
                compound_tuples_aux += 1

            # Verificar si la muestra ya existe
            existing_sample = session.query(Samples).filter_by(
                station_id=StationID,
                compound_cas=row[5].value,
                component_rt=row[0].value,
                library_rt=library_rt_value,  # row[1].value
                match_factor=row[3].value,
                sample_date=dt
            ).first()

            if not existing_sample:
                session.add(sample)
                sample_tuples_aux += 1

    compound_tuples += compound_tuples_aux
    sample_tuples += sample_tuples_aux

# Guardar los cambios en la base de datos (Esto estaba debajo del first() de existing_sample)
session.commit()

print("\nAll tuples were successfully inserted into the database.")
print(f"{compound_tuples} tuples inserted into Compounds")
print(f"{sample_tuples} tuples inserted into Samples")
