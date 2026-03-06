import os
import subprocess
import json
import sqlite3
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import geopandas as gpd
import altair as alt

from sqlalchemy import create_engine, CheckConstraint, String, Date, Numeric, Integer, ForeignKey
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker
from datetime import datetime, date
import openpyxl
import json

#------------------------------------------------------------
#------------------- CONFIGURACIÓN INICIAL ------------------
#------------------------------------------------------------
st.set_page_config(
    page_title="Pétrola Dashbaord",
    page_icon="📊",         # Puedes usar un emoji o una ruta a un ícono
    layout="wide",      # Opciones: "centered", "wide"
)

# CSS de estilos
st.markdown("""
    <style>
    /* --- PESTAÑAS --- */
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
    }

    .stTabs [data-baseweb="tab"] {
        padding: 10px 20px;
        border-radius: 4px 4px 0 0;
        background-color: #eaf7ff;
        color: black;
        transition: color 0.2s, background-color 0.2s;
    }
            
    .stTabs [data-baseweb="tab"]:hover {
        color: red;
    }

    .stTabs [aria-selected="true"] {
        background-color: #3498db;
        color: white;
    }

    body[data-theme="dark"] .stTabs [data-baseweb="tab"] {
        color: black;
    }


    /* --- MÉTRICAS --- */
    .metric-box {
        background-color: #eaf7ff;
        border-radius: 8px;
        padding: 15px;
        text-align: center;
        margin-bottom: 15px;
    }

    .metric-value {
        font-size: 1.8rem;
        font-weight: bold;
        color: #3498db;
    }

    .metric-label {
        font-size: 0.9rem;
        color: #7f8c8d;
    }

    /* --- FILTROS --- */
    .filter-container {
        background-color: #eaf7ff;
        border-radius: 10px;
        padding: 15px;
        margin-bottom: 20px;
    }

    .filter-title {
        font-weight: bold;
        margin-bottom: 10px;
        color: #2c3e50;
    }

    /* --- MULTISELECT --- */
    .stMultiSelect [data-baseweb="tag"] {
        background-color: #3498db !important;
        color: white !important;
        border: none !important;
    }
    .stMultiSelect div[data-baseweb="select"] span {
        font-size: 14px !important;  /* Ajusta este tamaño según necesites */
    }
""", unsafe_allow_html=True)




#------------------------------------------------------------
#------------------- VARIABLES GLOBALES ---------------------
#------------------------------------------------------------
database_path = "Database/Petrola.db"
database_url = f"sqlite:///{database_path}"

# ----------------------------------------------------------
# ----------------- FUNCIONES DASHBOARD --------------------
# ----------------------------------------------------------


def login():
    # Función para realizar login, comprueba la contraseña y cambia las variables de sesión de forma correspondiente
    username = st.session_state["input_user"]
    password = st.session_state["input_pass"]
    if USER_CREDENTIALS.get(username) == password:
        st.session_state.logged_in = True
        st.session_state.username = username
        st.session_state.login_error = False
    else:
        st.session_state.login_error = True


@st.cache_data(ttl=3600)
def cargar_dataframe_desde_db(db_path):
    # Esta función carga todas las tablas de la base de datos y las une en un solo dataframe para realizar los filtros y graficas mas facilmente
    # Esta función solo se ejecuta una vez cada hora o cuando se reinicie desde terminal el dashboard, para evitar consultas frecuentes irrelevantes con los mismos datos -> mejora velocidad de respuesta

    # Comprobamos que la base de datos existe y realizamos la conexión
    if not os.path.exists(db_path):
        st.error(f"No se encontró la base de datos en: {db_path}")
        return None
    conn = sqlite3.connect(db_path)

    # Obtenemos los datos de las tres tablas (samples, stations y compounds) en formato dataframe
    df_muestras = pd.read_sql_query("SELECT * FROM samples", conn)
    df_stations = pd.read_sql_query("SELECT * FROM stations", conn)
    df_compounds = pd.read_sql_query("SELECT * FROM compounds", conn)

    # Cerramos la conexión con la base de datos
    conn.close()

    # Unimos los tres dataframes por sus foreign keys
    df_muestras_stations = df_muestras.merge(df_stations, on='station_id', how='left')
    df_petrola = df_muestras_stations.merge(df_compounds, left_on='compound_cas', right_on='cas', how='left')

    # Convertimos las fechas para el filtrado por tiempo usando datetime
    df_petrola["sample_date"] = pd.to_datetime(df_petrola["sample_date"])
    return df_petrola

@st.cache_data(ttl=3600)
def generar_diccionario_de_colores_de_grupo(df_petrola):
    # Crea un diccionario de colores para cada grupo, de forma que se pueda usar el mismo color para cada grupo en diferentes tablas.
    # Al igual que la anterior, solo se ejecuta una vez cada hora para mejorar rendimiento
    # El diccionario creado tiene el formato {'nombre_grupo' : 'color'}

    grupos = df_petrola['group'].unique().tolist()
    diccionario_grupos_colores = {}
    lista_colores = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
    '#393b79', '#637939', '#8c6d31', '#843c39', '#7b4173',
    '#5254a3', '#9c9ede', '#17becf', '#9edae5', '#e6550d',
    '#ffbb78', '#98df8a', '#ff9896', '#c5b0d5', '#c49c94',
    '#f7b6d2', '#c7c7c7', '#dbdb8d', '#aec7e8', '#8c564b',
    '#bd9e39', '#e7ba52', '#d6616b', '#e7969c', '#7b4173',
    '#ce6dbd', '#de9ed6', '#6b6ecf', '#9c9ede', '#393b79',
    '#cedb9c', '#8ca252', '#e7cb94', '#bd9e39', '#ad494a',
    '#a55194', '#6b6ecf', '#637939', '#d62728', '#1f77b4'
]

    for i, grupo in enumerate(grupos):
        diccionario_grupos_colores[grupo] = lista_colores[i]

    return diccionario_grupos_colores


@st.cache_data(ttl=3600)
def generar_diccionario_de_colores_de_estacion(df_petrola):
    # Función idéntica a la anterior, pero asigna colores a las estaciones. Para que el grafico de barras y mapa tengan los mismos colores
    estaciones = df_petrola['station_id'].unique().tolist()
    diccionario_estaciones_colores = {}
    lista_colores = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
    '#393b79', '#637939', '#8c6d31', '#843c39', '#7b4173',
    '#5254a3', '#9c9ede', '#17becf', '#9edae5', '#e6550d',
    '#ffbb78', '#98df8a', '#ff9896', '#c5b0d5', '#c49c94',
    '#f7b6d2', '#c7c7c7', '#dbdb8d', '#aec7e8', '#8c564b',
    '#bd9e39', '#e7ba52', '#d6616b', '#e7969c', '#7b4173',
    '#ce6dbd', '#de9ed6', '#6b6ecf', '#9c9ede', '#393b79',
    '#cedb9c', '#8ca252', '#e7cb94', '#bd9e39', '#ad494a',
    '#a55194', '#6b6ecf', '#637939', '#d62728', '#1f77b4'
]

    for i, estacion in enumerate(estaciones):
        diccionario_estaciones_colores[estacion] = lista_colores[i]

    return diccionario_estaciones_colores

def plot_evolution_over_time(df_filtrado, orden_periodo, filtros, diccionario_colores):
    # Función que muestra la evolución de detección de muestras según los filtros seleccionados (Compuestos, grupos, estaciones y tiempo)

    # Se obtienen los filtros seleccionados
    compuesto_sel = filtros["compuestos"]
    familia_sel = filtros["familias"]
    estacion_sel = filtros["estaciones"]
    tipo_visualizacion = filtros["modo_estacion"]
    tipo_estacion = filtros["tipo_estacion"]

    # Se crea la figura
    fig = go.Figure()

    # Se agregan por compuestos, grupos o estaciones segun los filtros seleccionados.
    # La lógica se explica a continuación (0 para no seleccionados y 1 para seleccionados):
    #   Compuesto (0) - Grupo (0) -> Se muestran los 8 grupos mas frecuentes
    #   Compuesto (0) - Grupo (1) -> Se muestra la evolución de grupos seleccionados en la gráfica
    #   Compuesto (1) - Grupo (0) -> Se muestra la evolución de compuestos seleccionados en la gráfica
    #   Compuesto (1) - Grupo (1) -> Se muestra la evolución de compuestos seleccionados en la gráfica
    if tipo_visualizacion == "Grupo/Compuesto":
        if compuesto_sel:
            frecuencia = df_filtrado.groupby(['name', 'periodo']).size().unstack(fill_value=0)
            titulo = 'Compuesto'
            leyenda = titulo
            indices = compuesto_sel
        else:
            if not familia_sel or len(familia_sel) == 0:
                top_familias = df_filtrado['group'].value_counts().nlargest(8).index.tolist()
            else:
                top_familias = familia_sel

            df_limite_8_familias = df_filtrado[df_filtrado['group'].isin(top_familias)]
            frecuencia = df_limite_8_familias.groupby(['group', 'periodo']).size().unstack(fill_value=0)
            titulo = 'Grupo'
            leyenda = titulo
            indices = frecuencia.index.tolist()

        frecuencia = frecuencia.reindex(columns=orden_periodo, fill_value=0)

        for item in indices:
            if item in frecuencia.index:
                color = diccionario_colores.get(item, None) if titulo == 'Grupo' else None
                fig.add_trace(go.Scatter(
                    x=frecuencia.columns,
                    y=frecuencia.loc[item],
                    mode='lines+markers',
                    name=str(item),
                    marker=dict(size=8, color=color) if color else dict(size=8)
                ))
    elif tipo_visualizacion == "Estación":
        if compuesto_sel:
            if len(compuesto_sel) > 4:
                return
            
            if len(estacion_sel) > 4:
                return
            if len(estacion_sel) == 0 or estacion_sel == None:
                return
            
            titulo = 'Compuesto'
            leyenda = 'Compuesto - Estación'
            df_individual = df_filtrado[df_filtrado['name'].isin(compuesto_sel)]

            for compuesto in compuesto_sel:
                df_comp = df_individual[df_individual['name'] == compuesto]

                estaciones_unicas = df_comp['station_id'].unique()

                for estacion in estaciones_unicas:
                    df_est = df_comp[df_comp['station_id'] == estacion]
                    frecuencia = df_est.groupby('periodo').size().reindex(orden_periodo, fill_value=0)

                    fig.add_trace(go.Scatter(
                        x=frecuencia.index,
                        y=frecuencia.values,
                        mode='lines+markers',
                        name=f'{compuesto} - {estacion}',
                        marker=dict(size=8)
                    ))
        elif familia_sel:
            if len(familia_sel) > 4:
                return
            if len(estacion_sel) > 4:
                return
            if len(estacion_sel) == 0 or estacion_sel == None:
                return
            
            titulo = 'Grupo'
            leyenda = 'Grupo - Estación'
            df_individual = df_filtrado[df_filtrado['group'].isin(familia_sel)]

            for familia in familia_sel:
                df_comp = df_individual[df_individual['group'] == familia]

                estaciones_unicas = df_comp['station_id'].unique()

                for estacion in estaciones_unicas:
                    df_est = df_comp[df_comp['station_id'] == estacion]
                    frecuencia = df_est.groupby('periodo').size().reindex(orden_periodo, fill_value=0)

                    fig.add_trace(go.Scatter(
                        x=frecuencia.index,
                        y=frecuencia.values,
                        mode='lines+markers',
                        name=f'{familia} - {estacion}',
                        marker=dict(size=8)
                    ))
        else:
            return

    # Creación de un titulo dinamico, dependiendo de si se han seleccionado estaciones, grupos o compuestos.
    if estacion_sel:
        if len(estacion_sel) > 5:
            estaciones_str = ", ".join(str(e) for e in estacion_sel[:5]) + "..."
        else:
            estaciones_str = ", ".join(str(e) for e in estacion_sel)
        title = f'Evolución de detección de {titulo}s - Estación(es) {estaciones_str}'
    else:
        if tipo_estacion:
            tipos = " ,".join(str(tipo) for tipo in tipo_estacion)
            grupos_o_todas = f"Estaciones de tipo: {tipos}"
        else:
            grupos_o_todas = "Todas las Estaciones"

        title = f'Evolución de detección de {titulo}s - {grupos_o_todas}'


    # Se eliminan el nombre de los ejes, para que la gráfica sea más grande y se vea mejor
    fig.update_layout(
        title=title,
        xaxis_title="",
        yaxis_title="",
        xaxis=dict(tickangle=45),
        legend_title_text=leyenda,
        showlegend=True,
        height=620
    )

    # Se dibuja la gráfica en streamlit usando todo el ancho disponible
    st.plotly_chart(fig, use_container_width=True)


def plot_boxplot_match_factor(df_filtrado, familia_sel, diccionario_colores_grupos, valor_barras_horizontales, mostrar_quesito):
    # Función que genera una gráfica de boxplots, dependiendo de si se han elegido grupos o compuestos.
    # Sigue la misma lógica que la función anterior.
    
    # Originalmente se añadió un gráfico de quesito (ahora comentado). Ahora siempre se muestran 18 si no se ha seleccionado ninguno.
    if mostrar_quesito:
        numero_grupos_a_mostrar = 8
    else:
        if len(df_filtrado['group'].unique()) < 18:
            numero_grupos_a_mostrar = len(df_filtrado['group'].unique())
        else:
            numero_grupos_a_mostrar = 18
    
    if not familia_sel:
        top_familias = df_filtrado['group'].value_counts().nlargest(numero_grupos_a_mostrar).index.tolist()
    else:
        top_familias = familia_sel

    # Se filtra el dataframe por los grupos seleccionados
    df_boxplot = df_filtrado[df_filtrado['group'].isin(top_familias)]
    if df_boxplot.empty:
        st.write("No hay datos para mostrar.")
        return

    # Se crea el boxplot con los colores del diccionario de grupos
    fig = px.box(
        df_boxplot,
        x='group',
        y='match_factor',
        color='group',
        title='Distribución de Match Factor por Grupo',
        color_discrete_map=diccionario_colores_grupos,
        height=600
    )

    # Se elimina el nombre de los ejes para hacer la figura más grande
    fig.update_layout(showlegend=False, xaxis_title='', yaxis_title='')
    fig.update_yaxes(range=[70,100])

    # Se añaden dos lineas horizontales que muestran el rango de match factor seleccionado.
    fig.add_hline(y=valor_barras_horizontales[0], line_dash="dash", line_color="red")
    fig.add_hline(y=valor_barras_horizontales[1], line_dash="dash", line_color="red")

    # Se dibuja el gráfico en streamlit, usando todo el ancho disponible.
    st.plotly_chart(fig, use_container_width=True)
def plot_top_grupos(df_filtrado, diccionario_colores):
    # Gráfica de quesito mostrado originalmente.

    # Se calculan las frecuencias de cada grupo
    grupos_freq = df_filtrado["group"].value_counts()
        
    # Si existen más de 8 grupos diferentes, se muestran los 8 más frecuentes y se agrupan el resto en 'Otros'. Por temas de espacio en la visualización.
    if len(grupos_freq) > 8:
        top8 = grupos_freq[:8]
        otros = grupos_freq[8:].sum()
        top8["Otros"] = otros
    else:
        top8 = grupos_freq

    # Se calculan los compuestos más frecuentes de cada grupo.
    top_compuestos = (
        df_filtrado.groupby(["group", "name"])
        .size()
        .reset_index(name='frecuencia')
        .sort_values(['group', 'frecuencia'], ascending=[True, False])
    )

    # Se seleccionan los 3 más frecuentes de cada grupo, para la información del hover.
    top3_por_grupo = (
        top_compuestos
        .groupby("group")
        .head(3)
        .groupby("group")["name"]
        .apply(lambda x: " | ".join(x))
    )

    # Se crea el dataframe con los grupos que se van a mostrar, su frecuencia y los 3 compuestos más frecuentes de cada grupo.
    df_top8 = top8.reset_index()
    df_top8.columns = ['Grupo', 'Frecuencia']
    df_top8["Top compuestos"] = df_top8["Grupo"].map(top3_por_grupo)
    df_top8["Top compuestos"] = df_top8["Top compuestos"].fillna("No disponible")

    # Se crear el gráfico de quesito con los colores del diccionario
    fig = px.pie(
        df_top8,
        names='Grupo',
        values='Frecuencia',
        hover_data=['Top compuestos'],
        color='Grupo',
        color_discrete_map=diccionario_colores
    )
    
    # Se añade la información que se mostrará al hacer hover sobre cada porción. Nombre, NºMuestras y compuestos más frecuentes.
    fig.update_traces(
        textinfo='percent+label',
        hovertemplate="<b>%{label}</b><br>Nº muestras: %{value}<br>Top compuestos: %{customdata[0]}<extra></extra>"
    )
    
    # Se ajustan los margenes y proporciones para el espacio disponible y se muestra
    fig.update_layout(showlegend=False, height= 570, width= 570, title="Proporción de Grupos según detecciones",margin=dict(t=65, b=50, l=0, r=0),)
    st.plotly_chart(fig, use_container_width=True)


def plot_station_map_plotly(df_filtrado, estaciones_color):
    # Función que crea un mapa de la laguna con los limites hidrológicos y marca las estaciones como puntos en el mapa.

    # Se leen los límites y se convierten a WGS84
    bounds = gpd.read_file('Geolocalización Data/limite_hidrogeologico.shp')
    bounds_wsg84 = bounds.to_crs('EPSG:4326')
    bounds_geojson = bounds_wsg84.__geo_interface__

   
    # Se procesan las coordeandas de las estaciones para generar los marcadores
    estaciones = (
        df_filtrado[['station_id', 'x', 'y', 'st_type', 'geology']]
        .dropna(subset=['x', 'y'])
        .drop_duplicates(subset=['station_id'])
        .set_index('station_id')[['x', 'y', 'st_type', 'geology']]
    )

    # Se crea un geopandas dataframe con las estaciones y se pasa al sistema necesario para mostrar en la gráfica.
    geo_pts = gpd.points_from_xy(estaciones['x'], estaciones['y'], crs='EPSG:25830')
    stations_gdf = gpd.GeoDataFrame(estaciones, geometry=geo_pts)
    stations_wsg84 = stations_gdf.to_crs('EPSG:4326')


    # Se calcula el número de muestras por estación, para la información de hover
    num_muestras = df_filtrado.groupby('station_id').size().rename('muestras')
    stations_wsg84 = stations_wsg84.join(num_muestras, on='station_id')


    # Se crea el dataframe con cada estación, los compuestos y su numero de detecciones.
    compuestos_por_estacion = (
        df_filtrado.groupby(['station_id', 'name'])
        .size()
        .reset_index(name='cuenta')
    )

    # Se itera por cada estación, creando la inforamción de hover (ID, NºMuestras, Geologia y Compuestos) y añadiendola al diccionario.
    hover_data = {}
    for station_id, grp in compuestos_por_estacion.groupby('station_id'):
        grp_sorted = grp.sort_values('cuenta', ascending=False)
        top5 = grp_sorted.head(5)
        otros = grp_sorted['cuenta'].iloc[5:].sum()
        lines = [f"{row['name']}: {row['cuenta']}" for _, row in top5.iterrows()]
        if otros > 0:
            lines.append(f"Otros: {otros}")
        geol = stations_wsg84.loc[station_id, 'geology']
        total = stations_wsg84.loc[station_id, 'muestras']
        texto = (
            f"<b>Estación:</b> {station_id}<br>"
            f"<b>Muestras totales:</b> {total}<br>"
            f"<b>Geología:</b> {geol}<br>"
            f"<b>Compuestos:</b><br>" + "<br>".join(lines)
        )
        hover_data[station_id] = texto

    # Variables para mostrar las estaciones en el gráfico Plotly
    lats = stations_wsg84.geometry.y.tolist()
    lons = stations_wsg84.geometry.x.tolist()
    station_ids = stations_wsg84.index.tolist()
    hover_texts = [hover_data[st] for st in station_ids]

    # Se crea la figura
    fig = go.Figure()

    # Se añade la linea de limite de la laguna
    for feature in bounds_geojson['features']:
        coords = feature['geometry']['coordinates'][0]
        lon_coords = [pt[0] for pt in coords]
        lat_coords = [pt[1] for pt in coords]
        fig.add_trace(go.Scattermap(
            lon=lon_coords,
            lat=lat_coords,
            mode='lines',
            line=dict(width=2, color='blue'),
            fill='toself',
            fillcolor='rgba(0,0,255,0.1)',
            hoverinfo='none',
            showlegend=False
        ))

    # Se denfinen los colores para cada marcador (estación) segun el diccionario de colores y se dibujan en el mapa
    colores = [estaciones_color.get(est, '#cccccc') for est in station_ids]
    fig.add_trace(go.Scattermap(
        lon=lons,
        lat=lats,
        mode='markers',
        marker=go.scattermap.Marker(
            size=16,
            color=colores, 
            opacity=1
        ),
        text=hover_texts,
        hoverinfo='text',
        showlegend=False
    ))

    # Se personalizan los valores iniciales del mapa y su tamaño    
    fig.update_layout(
        map=dict(
            style='satellite',
            center=dict(lat=38.84, lon=-1.5609350),
            zoom=13
        ),
        margin=dict(l=0, r=0, t=0, b=0),
        height=585
    )

    # Se dibuja la gráfica en Streamlit
    st.plotly_chart(fig, use_container_width=True)

def obtener_estacion(fecha):
    # Esta función toma una fecha y devuelve la estación del año en la que se encuentra dicha fecha.
    mes = fecha.month
    if mes in [12, 1, 2]:
        return 'Invierno'
    elif mes in [3, 4, 5]:
        return 'Primavera'
    elif mes in [6, 7, 8]:
        return 'Verano'
    else:
        return 'Otoño'
def aplicar_filtros(df, filtros):
    # Función que toma el Dataframe original (sin filtrar) y aplica los filtros que se hayan seleccionado

    # En primer lugar se crea una copia del dataframe original, no queremos modificarlo.
    df_filtrado = df.copy()

    # filtros es un diccionario que contiene todos los filtros y su valor.
    # El proceso de filtrado consiste en verificar si el filtro tiene un valor seleccionado y filtrar por dicho valor.
    if filtros["compuestos"]:
        df_filtrado = df_filtrado[df_filtrado["name"].isin(filtros["compuestos"])]

    if filtros["familias"]:
        df_filtrado = df_filtrado[df_filtrado["group"].isin(filtros["familias"])]

    # De momento el rango de match factor no filtra el dataframe, solo mueve las lineas horizontales de la gráfica.
    # Se puede descomentar este codigo para añadir la funcionalidad de filtrar el dataframe por match_factor
    #df_filtrado = df_filtrado[
    #    (df_filtrado["match_factor"] >= match_factor_sel[0]) &
    #    (df_filtrado["match_factor"] <= match_factor_sel[1])
    #    ]

    if filtros["estaciones"]:
        df_filtrado = df_filtrado[df_filtrado["station_id"].isin(filtros["estaciones"])]

    if filtros["tipo_estacion"]:
        df_filtrado = df_filtrado[df_filtrado["st_type"].isin(filtros["tipo_estacion"])]

    if filtros["tipo_tiempo"] == "Intervalo":
        inicio, fin = filtros["rango_fechas"]
        df_filtrado = df_filtrado[
            (df_filtrado["sample_date"] >= pd.to_datetime(inicio)) &
            (df_filtrado["sample_date"] <= pd.to_datetime(fin))
        ]

    # Periodo para gráficos
    if filtros["tipo_tiempo"] == "Mensual":
        df_filtrado["periodo"] = df_filtrado["sample_date"].dt.strftime('%b')
        orden = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    elif filtros["tipo_tiempo"] == "Estacional":
        df_filtrado["periodo"] = df_filtrado["sample_date"].apply(obtener_estacion)
        orden = ['Primavera', 'Verano', 'Otoño', 'Invierno']
    else:
        df_filtrado["periodo"] = df_filtrado["sample_date"].dt.to_period("M").dt.to_timestamp()
        orden = sorted(df_filtrado['periodo'].unique())

    return df_filtrado, orden

def sheet_is_valid(sheet):
    # Si el nombre de la página tiene menos de 8 caracteres, la página no es valida.
    return not len(sheet.title) < 8

def row_is_valid(row):
    # Si todos los valores de la fila son None o vacios, la fila no es válida
    return any(cell.value not in (None, "", '') for cell in row[:7])

def get_station_date(Sample_Name):
    # Extrae la fecha y la StationID del atributo Sample Name
    StationID = Sample_Name.split('_')[0]
    if len(Sample_Name.split('_')[1]) <= 2:
        day, month, year = Sample_Name.split('_')[1:4]
    else:
        day = Sample_Name.split('_')[1][0:2]
        month = Sample_Name.split('_')[1][2:4]
        year = '20' + Sample_Name.split('_')[1][4:6]
    if StationID == '2571b':
        StationID = '2571'
    return StationID, datetime(int(year), int(month), int(day)).date()
def insertar_nuevas_muestras(uploaded_file):
    # Función que añade nuevas muestras válidas a la bbdd, dado un archivo excel.
    engine = create_engine(database_url)
    Session = sessionmaker(bind=engine)
    session = Session()

    # ---------------------- CLASES ORM ------------------------------------
    class Base(DeclarativeBase):
        pass

    class Stations(Base):
        __tablename__ = "stations"
        station_id: Mapped[str] = mapped_column(String, primary_key=True)
        st_type: Mapped[str] = mapped_column(String(20), nullable=True)
        geology: Mapped[str] = mapped_column(String(20), nullable=True)
        x: Mapped[float] = mapped_column(Numeric(10, 3), nullable=False)
        y: Mapped[float] = mapped_column(Numeric(10, 3), nullable=False)

    class Compounds(Base):
        __tablename__ = "compounds"
        cas: Mapped[str] = mapped_column(String(20), primary_key=True)
        name: Mapped[str] = mapped_column(String(80), nullable=False)
        formula: Mapped[str] = mapped_column(String(20), nullable=False)
        group: Mapped[str] = mapped_column(String(30), nullable=True)

    class Samples(Base):
        __tablename__ = "samples"
        id: Mapped[int] = mapped_column(Integer, autoincrement=True, primary_key=True)
        station_id: Mapped[str] = mapped_column(ForeignKey("stations.station_id"), nullable=False)
        compound_cas: Mapped[str] = mapped_column(ForeignKey("compounds.cas"), nullable=False)
        component_rt: Mapped[float] = mapped_column(Numeric(17, 13), nullable=False)
        library_rt: Mapped[float] = mapped_column(Numeric(17, 13), nullable=True)
        match_factor: Mapped[float] = mapped_column(Numeric(16, 13), nullable=False)
        sample_date: Mapped[date] = mapped_column(Date, nullable=False)

        __table_args__ = (
            CheckConstraint("match_factor >= 0 AND match_factor <= 100", name="check_match_factor_range"),
        )

    #---------------SCRIPT DE INSERCCION---------------------------
    workbook = openpyxl.load_workbook(uploaded_file)
    compound_tuples_aux = 0
    sample_tuples_aux = 0
    total_rows = sum(sheet.max_row - 4 for sheet in workbook.worksheets)

    # Barra para informar del progreso al usuario desde el dashboard
    progress_bar = st.progress(0)

    i = 0
    # Se itera por cada página y se comprueba que es válida
    for sheet in workbook.worksheets:
        if not sheet_is_valid(sheet):
            continue

        # Se itera por cada fila y se comprueba que es válida
        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
            i += 1
            progress_bar.progress(i / total_rows)

            if not row_is_valid(row):
                continue

            try:
                try:
                    group_value = row[7].value if row[7].value not in (None, '') else 'Otros'
                except Exception as e:
                    group_value = 'Otros'

                compound = Compounds(
                    cas=row[5].value,
                    name=row[2].value,
                    formula=row[4].value,
                    group=group_value
                )

                library_rt_value = row[1].value if row[1].value != '' else None
                StationID, dt = get_station_date(row[6].value)

                # Se comprueba que los datos no existen ya en la base de datos
                existing_station = session.query(Stations).filter_by(station_id=StationID).first()
                if not existing_station:
                    continue
                existing_compound = session.query(Compounds).filter_by(cas=row[5].value).first()
                if not existing_compound:
                    session.add(compound)
                    compound_tuples_aux += 1
                existing_sample = session.query(Samples).filter_by(
                    station_id=StationID,
                    compound_cas=row[5].value,
                    component_rt=row[0].value,
                    library_rt=library_rt_value,
                    match_factor=row[3].value,
                    sample_date=dt
                ).first()

                # Si no existen, se añaden
                if not existing_sample:
                    sample = Samples(
                        station_id=StationID,
                        compound_cas=row[5].value,
                        component_rt=row[0].value,
                        library_rt=library_rt_value,
                        match_factor=row[3].value,
                        sample_date=dt
                    )
                    session.add(sample)
                    sample_tuples_aux += 1
            except Exception as e:
                st.error(f"Error procesando fila: {e}")
                continue

    # Guardamos los cambios y cerramos la conexion con la base de datos
    session.commit()
    session.close()
    engine.dispose()

    # Mostramos mensaje de finalización y ocultamos la barra de progreso
    progress_bar.empty()

    return compound_tuples_aux, sample_tuples_aux





# def pushear_bbdd(df_petrola, total):
#     # Función que hace un commit y puseha el archivo petrola.db, para actualizar la base de datos del repositorio de forma persistente.

#     # Información requerida para realizar el push al repositorio
#     file_path = "Database/Petrola.db"
#     github_token = os.environ["GITHUB_TOKEN"]
#     repo_url = "github.com/LagunaPetrolaDashboard/Dashboard-Laguna-Petrola-2025"
#     commit_message = f"Añadidas nuevas muestras a base de datos. Total Muestras: {len(df_petrola)+int(total)}"
#     authed_repo_url = f"https://{github_token}@{repo_url}"

#     try:
#         # Se asegura estar en la rama main para el push
#         subprocess.run(["git", "checkout", "-B", "main"], check=True)

#         # Se configura un usuario de Git para realizar el commit
#         subprocess.run(["git", "config", "user.name", "Usuario Externo Community Cloud"], check=True)
#         subprocess.run(["git", "config", "user.email", "streamlit@example.com"], check=True)

#         # Se hace añade el archivo al stage
#         subprocess.run(["git", "add", file_path], check=True)

#         # Se hace el commit de la base de datos con el mensaje predefinido
#         subprocess.run(["git", "commit", "-m", commit_message], check=True)

#         # Se verifica si remote existe, si no existe se crea
#         remotes = subprocess.check_output(["git", "remote"]).decode().split()
#         if "origin" not in remotes:
#             subprocess.run(["git", "remote", "add", "origin", authed_repo_url], check=True)
#         else:
#             subprocess.run(["git", "remote", "set-url", "origin", authed_repo_url], check=True)

#         # Se hace el push a la rama main
#         subprocess.run(["git", "push", "origin", "main"], check=True)

#     except subprocess.CalledProcessError as e:
#         print(f"Error durante el push: {e}")

#     except Exception as e:
#         print(f"Error general: {e}")





#------------------------------------------------------------
#------------------- AUTENTIFICACIÓN DE USUARIO--------------
#------------------------------------------------------------
#json_diccionario = os.environ["USER_CREDENTIALS"]

#USER_CREDENTIALS = json.loads(json_diccionario)
USER_CREDENTIALS = {"admin": "miTFG-2025", "LP2025": "Muestras_Pétrola25"}

# Estado de Sesion
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "username" not in st.session_state:
    st.session_state.username = ""
if "login_error" not in st.session_state:
    st.session_state.login_error = False


# Interfaz de login
if not st.session_state.logged_in:
    st.markdown("""
        <h1 style='text-align: center; 
                font-family: "Segoe UI", "Helvetica Neue", sans-serif; 
                font-size: 2.2em; 
                margin-bottom: 0.2em;'>
            Herramienta de Análisis de Muestras<br>Laguna de Pétrola
        </h1>
    """, unsafe_allow_html=True)
    st.markdown("---")

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.text_input("Usuario", key="input_user")
        st.text_input("Contraseña", type="password", key="input_pass")
        st.button("Acceder", use_container_width=True, on_click=login)

        if st.session_state.login_error:
            st.error("Usuario o contraseña incorrectos.")

    st.stop()





# --------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------ CONTENIDO DASHBOARD -----------------------------------------------
# --------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------

st.subheader("Análisis de muestras de agua – Laguna de Pétrola")
st.markdown("""
    <style>
    .block-container {
        padding-top: 1.6rem;
    }
    </style>
""", unsafe_allow_html=True)



# Se carga el dataframe, extrayendo los datos de la base de datos
df_petrola = cargar_dataframe_desde_db(database_path)

# Se inicializa el diccionario de filtros
filtros = {}

# Se generan los colores de los grupos
grupos_color = generar_diccionario_de_colores_de_grupo(df_petrola)

# Se generan los colores de las estaciones
estaciones_color = generar_diccionario_de_colores_de_estacion(df_petrola)

# Espacio reservado para situar las métricas en la parte superior del dashboard
metrics_placeholder = st.empty()

# Reserva un espacio fijo para que cuando se recargue el dataframe, no suban todos los elementos y vuelvan a bajar de rápidamente
metrics_placeholder.markdown(
    "<div style='height:115px'></div>", 
    unsafe_allow_html=True
)
st.text("")

# Parte inferior (Filtros)
with st.container():
    st.markdown("<hr style='border:1px solid #ccc; margin-top:2px; margin-bottom:2px;' />", unsafe_allow_html=True)
    compuestos_tab, stations_tab, time_tab, filtered_data_tab, add_data_tab = st.tabs(["🧪 Compuestos y Grupos", "📡 Estaciones", "📈 Evolución Temporal","📄 Datos Filtrados", "➕ Añadir nuevos datos"])
    
    with compuestos_tab:

        # ---------------------------------------------------------------------
        # ----------------PESTAÑA 1: Compuestos y Grupos (Filtros) ------------
        # ---------------------------------------------------------------------

        filtros_match_bar, resto = st.columns([1.5, 4.65])

        with filtros_match_bar:
            filtros_match_bar.text("")
            filtros_match_bar.text("")


            st.markdown("""<div style='font-weight: 900; display: block; margin: 0; padding: 0;'>Selección de Compuestos</div>""",unsafe_allow_html=True)
            
            
            # Familias / Grupos
            familias = sorted(df_petrola["group"].dropna().unique())
            familia_sel = st.multiselect(
                "Grupo(s)", 
                familias, 
                default=None,
                placeholder="Todos"
            )
            filtros['familias'] = familia_sel if familia_sel else []
            

            # Compuestos
            if familia_sel:
                df_compuestos_por_grupo = df_petrola[df_petrola['group'].isin(familia_sel)]
                nombre_grupos = ", ".join(familia_sel)
                nombre_grupos = "Pertenecientes a: "+nombre_grupos
            else:
                df_compuestos_por_grupo = df_petrola
                nombre_grupos = 'Todos'

            compuestos = sorted({f"{name} [{cas}]" for name, cas in zip(df_compuestos_por_grupo['name'], df_compuestos_por_grupo['cas'])})
            compuesto_sel = st.multiselect("Compuesto(s)", compuestos, default=None, placeholder=nombre_grupos, help="Selecciona uno o varios compuestos por nombre o CAS. Si se ha seleccionado un grupo, solo serán seleccionables los compuestos pertenecientes a ese grupo.")
            nombre_compuesto_sel = [compuesto.split(' [')[0] for compuesto in compuesto_sel]
            filtros['compuestos'] = nombre_compuesto_sel

            filtros_match_bar.text("")
            filtros_match_bar.text("")

            st.markdown("<span style='font-weight: 900;'>Rango de Match Factor</span>", unsafe_allow_html=True)
            min_mf = float(df_petrola["match_factor"].min())
            max_mf = float(df_petrola["match_factor"].max())
            match_factor_sel = st.slider(
                "Selecciona rango Match Factor",
                min_value=min_mf,
                max_value=max_mf,
                value=(min_mf, max_mf),
                step=0.1
            )
            filtros['match_factor'] = (min_mf, max_mf)
            
            filtros_match_bar.text("")
            filtros_match_bar.text("")
            # mostrar_quesito = st.checkbox("Mostrar Gráfica de Proporciones", value=False)
        
        with resto:
            #if mostrar_quesito:
            #    vacio_match_bar, grafica_match_bar, grafica_quesito = st.columns([0.15, 2.5,2])
            #else:
            #    vacio_match_bar, grafica_match_bar, grafica_quesito = st.columns([0.15,3,0.000000001])
            vacio_match_bar, grafica_match_bar, grafica_quesito = st.columns([0.15,3,0.000000001])


        with vacio_match_bar:
            pass
    
    with stations_tab:

        # ---------------------------------------------------------------------
        # -------------------PESTAÑA 2: Estaciones (Filtros) ------------------
        # ---------------------------------------------------------------------

        filtros_stations_tab, vacio_stations_tab, grafica_stations_tab = st.columns([1,0.1, 2])
        with filtros_stations_tab:
            filtros_stations_tab.text("")
            st.markdown("<span style='font-weight: 900;'>Selección de estación</span>", unsafe_allow_html=True)

             # Tipo de estación
            tipos_estacion = sorted(df_petrola["st_type"].dropna().unique())
            tipo_estacion_sel = st.multiselect("Tipos de estación", tipos_estacion, default=None, placeholder='Todos')
            filtros['tipo_estacion'] = tipo_estacion_sel

            # Estaciones
            estaciones = sorted(df_petrola["station_id"].dropna().unique())
            estacion_sel = st.multiselect("Estación(es)", estaciones, default=None, placeholder='Todos')
            filtros['estaciones'] = estacion_sel

           

        with vacio_stations_tab:
            pass


    with time_tab:
        # ---------------------------------------------------------------------
        # -----------PESTAÑA 3: Evolición Temporal (Filtros) ------------------
        # ---------------------------------------------------------------------

        filtros_time_bar, vacio_time_bar, grafica_time_bar = st.columns([1,0.1, 3])
        with filtros_time_bar:
            filtros_time_bar.text("")
            filtros_time_bar.text("")
            st.markdown("<span style='font-weight: 900;'>Selección de Tiempo</span>", unsafe_allow_html=True)

            # Tipo de filtro temporal, si se elige Intervaloo, aparece la opción de elegir rango de fechas
            tipo_tiempo = st.selectbox("Filtrar por periodo de tiempo:", ["Mensual", "Estacional", "Intervalo"])
            filtros['tipo_tiempo'] = tipo_tiempo


            # La fecha de elección estará entre la muestra mas antigua y mas reciente detectada
            min_fecha = df_petrola["sample_date"].min()
            max_fecha = df_petrola["sample_date"].max()

            if tipo_tiempo == "Intervalo":
                rango_fechas_sel = st.date_input(
                    "Selecciona rango de fechas",
                    value=(min_fecha, max_fecha),
                    min_value=min_fecha,
                    max_value=max_fecha
                )
                filtros['rango_fechas'] = rango_fechas_sel if len(rango_fechas_sel) == 2 else (min_fecha, max_fecha)
            else:
                filtros['rango_fechas'] = (min_fecha, max_fecha)

            # Agregar por Grupo/Compuesto o Estación
            filtros_time_bar.text("")
            filtros_time_bar.text("")
            st.markdown("<span style='font-weight: 900;'>Agregar por </span>", unsafe_allow_html=True)
            modo_estacion = st.selectbox("Tipo de agregación:", ["Grupo/Compuesto", "Estación"])
            filtros['modo_estacion'] = modo_estacion

            if modo_estacion == "Estación" and (len(nombre_compuesto_sel) > 4 or len(familia_sel) > 4 or len(estacion_sel) > 4):
                st.warning(
                    "⚠️ Atención: El modo Estación solo permite seleccionar un máximo de 4 compuestos o grupos y 4 estaciones.")
                
            if modo_estacion == "Estación" and ( (len(compuesto_sel) < 1 and len(familia_sel)<1)   or len(estacion_sel) < 1) :
                st.warning(
                    "⚠️ Atención: Debes seleccionar al menos 1 compuesto o grupo y 1 estación para visualizar la gráfica en modo Agregado por Estación.")

        with vacio_time_bar:
            st.markdown("")



    with add_data_tab:
        # ---------------------------------------------------------------------
        # ------------------ PESTAÑA 5: Añadir nuevos datos -------------------
        # ---------------------------------------------------------------------

        informacion, vacio_add, dataframe_muestra_inserccion = st.columns([1,0.1, 1.5])
        with dataframe_muestra_inserccion:
            dataframe_muestra_inserccion.text("")
            st.markdown("<span style='font-weight: 900;'>Archivo con estructura válida para la inserción</span>", unsafe_allow_html=True)
            st.image("Images/Estructura_excel_referencia_para_insertado_de_nuevos_datos.JPG")

        with informacion:
            #st.subheader("Importar Datos")
            informacion.text("")
            st.markdown("<span style='font-weight: 900;'>Importar nuevos datos</span>", unsafe_allow_html=True)

            uploaded_files = st.file_uploader(
                "Subir uno o varios archivos Excel",
                type=["xls", "xlsx"],
                accept_multiple_files=True,
                key="excel_uploader"
            )

            if st.button("Iniciar proceso de importación", key="import_button"):
                if uploaded_files:
                    total_compounds = 0
                    total_samples = 0
                    for uploaded_file in uploaded_files:
                        st.write(f"Procesando archivo: {uploaded_file.name}")
                        compounds, samples = insertar_nuevas_muestras(uploaded_file)
                        total_compounds += compounds
                        total_samples += samples
                    st.success(
                        f"Importación completada con éxito. \nCompuestos insertados: {total_compounds} \nMuestras insertadas: {total_samples}. "
                        "\nPor favor, recargue el panel para visualizar los nuevos datos."
                    )
                    if total_compounds != 0 or total_samples != 0:
                        #pushear_bbdd(df_petrola, total_samples)
                        st.cache_data.clear()
                else:
                    st.warning("Por favor, sube al menos un archivo Excel en el formato correcto.")
            st.info(
                "**Requerimientos del archivo Excel:**\n"
                "- Solo se aceptan archivos Excel, en formato `.xlsx` o `.xls`.\n"
                "- Las 4 primeras filas de cada página serán ignoradas; pueden estar vacías.\n"
                "- El orden de las variables debe ser exactamente el mostrado en la imagen de referencia.\n"
                "- La columna `Compound Group` puede no existir, en ese caso se clasificacarán los compuestos nuevos como Otros.\n\n"

                "**Detalles relevantes del proceso de inserción:**\n"
                "- Las muestras duplicadas no se añadirán a la base de datos.\n"
                "- La fecha de la muestra será extraída de la variable `Sample Name`.\n"
                "- No interrumpir el proceso de inserción.\n"
            )

        with vacio_add:
            pass


# Se filtra el Dataframe usando los filtros seleccionados
df_filtrado, orden = aplicar_filtros(df_petrola, filtros)


# Se muestran las estadisticas generadas una vez filtrado el dataframe
with metrics_placeholder.container():
    col1, col2, col3, col4 = st.columns(4)
    col1.markdown(
        f"<div class='metric-box'><div class='metric-value'>{len(df_filtrado)}  💧</div><div class='metric-label'>Muestras</div></div>",
        unsafe_allow_html=True
    )
    col2.markdown(
        f"<div class='metric-box'><div class='metric-value'>{len(df_filtrado['station_id'].unique())}  📡</div><div class='metric-label'>Estaciones</div></div>",
        unsafe_allow_html=True
    )
    col3.markdown(
        f"<div class='metric-box'><div class='metric-value'>{len(df_filtrado['name'].unique()) }  🧪</div><div class='metric-label'>Compuestos</div></div>",
        unsafe_allow_html=True
    )
    col4.markdown(
        f"<div class='metric-box'><div class='metric-value'>{df_filtrado['match_factor'].mean():.1f}  % </div><div class='metric-label'>Media Match Factor</div></div>",
        unsafe_allow_html=True
    )





with compuestos_tab:
    # ---------------------------------------------------------------------
    # ---------- PESTAÑA 1: Compuestos y Grupos (Gráficas) -----------------
    # ---------------------------------------------------------------------
    with grafica_match_bar:
        valor_barras_horizontales = [match_factor_sel[0], match_factor_sel[1]]
        mostrar_quesito = False # Se ha decido no mostrar esta grafica, se deja por si acaso se cambia de opinion
        plot_boxplot_match_factor(df_filtrado, filtros['familias'], grupos_color, valor_barras_horizontales, mostrar_quesito)

        
        
    with grafica_quesito:
        #if mostrar_quesito:
        #    grafica_quesito.text("")
        #    plot_top_grupos(df_filtrado, grupos_color)
        pass





with time_tab:
    # ---------------------------------------------------------------------
    # ---------- PESTAÑA 3: Evolución Temporal (Gráfica) ------------------
    # ---------------------------------------------------------------------
    with grafica_time_bar:
        plot_evolution_over_time(df_filtrado, orden, filtros, grupos_color)





with stations_tab:
    # ---------------------------------------------------------------------
    # -------------------PESTAÑA 2: Estaciones (Gráficas) ------------------
    # ---------------------------------------------------------------------
    with filtros_stations_tab:
        filtros_stations_tab.text("")
        st.markdown("<span style='font-weight: 900;'>Muestras detectadas por estación</span>", unsafe_allow_html=True)


        # Muestras por estación
        muestras_por_estacion = df_filtrado['station_id'].value_counts()


        df = muestras_por_estacion.reset_index()
        df.columns = ['station_id', 'n_muestras']
        estaciones_tipo = df_filtrado[['station_id', 'st_type']].drop_duplicates()
        df = df.merge(estaciones_tipo, on='station_id', how='left')
        df['color'] = df['station_id'].map(estaciones_color)

        chart = alt.Chart(df).mark_bar().encode(
            x=alt.X('station_id:N', title='Estación'),
            y=alt.Y('n_muestras:Q', axis=alt.Axis(title=None)),
            color=alt.Color('color:N', scale=None),
            tooltip=[
                alt.Tooltip('station_id:N', title='Estación'),
                alt.Tooltip('n_muestras:Q', title='Nº de muestras'),
                alt.Tooltip('st_type:N', title='Tipo')
            ]

        ).properties(
            height=390
        )

        st.altair_chart(chart, use_container_width=True)

    with grafica_stations_tab:
        grafica_stations_tab.text("")
        st.markdown("<span style='font-weight: 900;'>Localización de Estaciones y Compuestos</span>",
                    unsafe_allow_html=True)
        plot_station_map_plotly(df_filtrado, estaciones_color)





with filtered_data_tab:
    # ---------------------------------------------------------------------
    # --------- PESTAÑA 4: Datos Filtrados (Filtros y Tabla) --------------
    # ---------------------------------------------------------------------
    filtered_data_tab.text("")
    n = st.selectbox("Número de muestras a mostrar", ["100","1000","10000","Todas"])

    # Omitimos mostrar estas columnas ya que no aportan demasiada información y solo hacen mas grande la tabla
    df_mostrar = df_filtrado.drop(columns=['id', 'periodo', 'x', 'y', 'compound_cas'])
    df_mostrar['sample_date'] = df_mostrar['sample_date'].dt.date
    n = len(df_mostrar) if n=="Todas" else int(n)

    # Mostramos la tabla sin restricciones de tamaño, se adaptará al espacio disponible
    st.dataframe(df_mostrar.head(n), height=540)
            
